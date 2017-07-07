import argparse
import random
import sys

import requests
from lxml import etree, html
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill


def prepare_session():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko)'
                      ' Chrome/56.0.2924.87 Safari/537.36',
        'Accept-Language': 'en-US;q=0.8,en;q=0.3',
        'Accept': 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    }
    http_session = requests.Session()
    http_session.headers.update(headers)
    return http_session


def fetch_url(url, session=None):
    session = session or prepare_session()
    try:
        webpage_data = session.get(url, timeout=30)
        webpage_data.raise_for_status()
    except requests.exceptions.ConnectionError:
        return None, 'Network problem while processing {}'.format(url)
    except requests.exceptions.Timeout:
        return None, 'Request times out while processing {}'.format(url)
    except requests.exceptions.TooManyRedirects:
        return None, 'Too many redirects at'.format(url)
    except requests.exceptions.HTTPError:
        return None, 'HTTP error occured while processing {}'.format(url)
    return webpage_data, None


def get_courses_list(xml_data, number_of_courses):
    sitemap = etree.fromstring(xml_data)
    my_namespace = {'default': sitemap.nsmap[None]}
    all_links = sitemap.xpath('//default:loc', namespaces=my_namespace)
    if number_of_courses > len(all_links):
        number_of_courses = len(all_links)
    return [link.text for link in random.sample(all_links, number_of_courses)]


def get_course_title(page_element):
    title_element = page_element.xpath('//h1[contains(concat(" ", @class, " "), "title")]')
    if title_element:
        return str(title_element[0].text_content())


def get_course_language(page_element):
    lang_element = page_element.xpath('//div[contains(concat(" ", @class, " "), "language-info")]/div')
    if lang_element:
        return str(lang_element[0].text_content())


def get_course_rating(page_element):
    rating_element = page_element.xpath('//div[contains(concat(" ", @class, " "), "ratings-text")]')
    if rating_element:
        text = str(rating_element[0].text_content())
        return ''.join(x for x in text if x in '0123456789.')


def get_course_duration(page_element):
    duration_element = page_element.xpath('//div[contains(concat(" ", @class, " "), "week-body")]')
    if duration_element:
        return len(duration_element)


def get_course_start_date(page_element):
    start_date = page_element.xpath('//div[contains(concat(" ", @class, " "), "startdate")]')
    if start_date:
        return str(start_date[0].text_content())


def crawl_courses_info(urls_list, session=None):
    courses_data = []
    total_url = len(urls_list)
    for url_num, url in enumerate(urls_list, start=1):
        print('Processing page {}/{} at: {}'.format(url_num, total_url, url))
        page_data, error = fetch_url(url, session=session)
        if error:
            print('[ERROR] {}'.format(error))
            continue
        course_page = html.fromstring(page_data.content.decode('utf-8'))
        course_info = {
            'title': get_course_title(course_page),
            'language': get_course_language(course_page),
            'start_date': get_course_start_date(course_page),
            'rating': get_course_rating(course_page),
            'duration': get_course_duration(course_page),
            'url': url
        }
        courses_data.append(course_info)
    return courses_data


def style_range(worksheet, cell_range, bg_color=None, align=None):
    """
    Apply styles to a range of cells
    https://openpyxl.readthedocs.io/en/default/styles.html#applying-styles

    :param worksheet: Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param bg_color: Background color in hex (e.g. ffffff or d7e4bc)
    :param align: An openpyxl Alignment object
    :return: None
    """
    bg_color = bg_color or 'ffffff'
    align = align or Alignment(horizontal='left', vertical='center', wrapText=True)
    bd = Side(style='thin', color='000000')
    cell_border = Border(left=bd, right=bd, top=bd, bottom=bd)
    for row in worksheet[cell_range]:
        fill_color = PatternFill('solid', fgColor=bg_color)
        for cell in row:
            cell.border = cell_border
            cell.fill = fill_color
            cell.alignment = align


def output_courses_info_to_xlsx(filepath, data_stream, worksheet_title='Courses information'):
    name_of_cols = ['№', 'Title', 'Languages', 'Rating', 'Start date', 'Duration (weeks)', 'URL']
    cols_sizes = {
        'A': 4,
        'B': 54,
        'C': 25,
        'D': 14,
        'E': 21,
        'F': 11,
        'G': 61
    }
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title
    for idx, name in enumerate(name_of_cols, start=1):
        worksheet.cell(row=1, column=idx, value=name)
    for name, size in cols_sizes.items():
        worksheet.column_dimensions[name].width = size
    style_range(worksheet, 'A1:G1', bg_color='d7e4bc')
    for idx, course_info in enumerate(data_stream, start=2):
        worksheet.cell(row=idx, column=1, value=idx - 1)
        worksheet.cell(row=idx, column=2, value=course_info['title'])
        worksheet.cell(row=idx, column=3, value=course_info['language'])
        worksheet.cell(row=idx, column=4, value=course_info['rating'])
        worksheet.cell(row=idx, column=5, value=course_info['start_date'])
        worksheet.cell(row=idx, column=6, value=course_info['duration'])
        worksheet.cell(row=idx, column=7, value=course_info['url'])
        style_range(worksheet, 'A{0}:G{0}'.format(idx))

    try:
        workbook.save(filename=filepath)
    except PermissionError:
        filepath = input('Unable to save to {}. Enter new filename'.format(filepath))
        workbook.save(filename=filepath)


def create_parser():
    argument_parser = argparse.ArgumentParser()
    argument_parser.add_argument('filepath', default='courses.xlsx', help='Save to', nargs='?')
    argument_parser.add_argument('number', default=20, type=int, help='How many courses to process', nargs='?')
    return argument_parser


def get_coursera_sitemap(session=None):
    xml, error = fetch_url('https://www.coursera.org/sitemap~www~courses.xml', session=session)
    if error:
        print('[ERROR] {}'.format(error))
        return None
    return xml.content


if __name__ == '__main__':
    parser = create_parser()
    namespace = parser.parse_args()

    download_session = prepare_session()
    sitemap_xml = get_coursera_sitemap(session=download_session)
    if sitemap_xml is None:
        sys.exit(1)

    links = get_courses_list(sitemap_xml, number_of_courses=namespace.number)
    info = crawl_courses_info(links, session=download_session)
    output_courses_info_to_xlsx(namespace.filepath, info)
